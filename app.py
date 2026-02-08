import io
import datetime
import pytz
import pandas as pd
import streamlit as st

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

TZ = "America/New_York"

# ---------------- helpers ----------------
def detect_col(df: pd.DataFrame, key: str):
    key = key.upper()
    for c in df.columns:
        if key in str(c).upper():
            return c
    return None

def style_header(ws):
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

def autosize(ws, cap=45, sample_rows=400):
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        max_len = 10
        for cell in col_cells[:sample_rows]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max_len + 2, cap)

def apply_route_colors(ws, colnames):
    yellow = PatternFill("solid", fgColor="FFF2CC")  # >30
    red = PatternFill("solid", fgColor="F8CBAD")     # >60
    purple = PatternFill("solid", fgColor="E4DFEC")  # no delivered

    min_idx = colnames.index("MinutesSinceLast") + 1
    flag_idx = colnames.index("StatusFlag") + 1
    rate_idx = colnames.index("CompletionRate") + 1

    for r in range(2, ws.max_row + 1):
        ws.cell(r, rate_idx).number_format = "0.00%"
        minutes = ws.cell(r, min_idx).value
        flag = ws.cell(r, flag_idx).value

        fill = None
        if flag == "NO_DELIVERED":
            fill = purple
        elif minutes is not None:
            try:
                m = float(minutes)
                if m > 60:
                    fill = red
                elif m > 30:
                    fill = yellow
            except:
                pass

        if fill:
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).fill = fill

def build_excel_bytes(raw_df: pd.DataFrame) -> bytes:
    """
    Input: åŽŸå§‹è¡¨ï¼ˆB=Route, J=Status, L=Timeï¼‰
    Output sheets:
      - RouteMonitor
      - Summary
      - Exceptions
      - 3pm check (>=3pm ET and <50%)
      - 6pm check (>=6pm ET and <80%)
      - Meta
    """
    df = raw_df.copy()

    # å›ºå®šåˆ—ä½ç½®ï¼šB/J/L
    route_col = df.columns[1]
    status_col = df.columns[9]
    time_col = df.columns[11]

    # è¯†åˆ«å­—æ®µï¼ˆå¯é€‰ï¼‰
    flee_col = detect_col(df, "FLEE")
    driver_col = detect_col(df, "DRIVER")

    df[time_col] = pd.to_datetime(df[time_col], errors="coerce")
    df["StatusU"] = df[status_col].astype(str).str.upper()

    tz = pytz.timezone(TZ)
    now_et = datetime.datetime.now(tz)                # for time checks
    now_et_naive = now_et.replace(tzinfo=None)        # for time diff calc

    rows = []
    for route, g in df.groupby(route_col, dropna=True):
        total = int(len(g))
        delivered_cnt = int((g["StatusU"] == "DELIVERED").sum())
        failed_cnt = int(g["StatusU"].str.contains("FAIL", na=False).sum())
        remaining = int(total - delivered_cnt - failed_cnt)
        completion = (delivered_cnt / total) if total else 0.0

        flee = g[flee_col].dropna().iloc[0] if flee_col and g[flee_col].notna().any() else None
        driver = g[driver_col].dropna().iloc[0] if driver_col and g[driver_col].notna().any() else None

        delivered_rows = g[(g["StatusU"] == "DELIVERED") & g[time_col].notna()]
        if delivered_rows.empty:
            first_del = None
            last_del = None
            minutes_since_last = None
            hours_since_first = None
            per_hour = None
            status_flag = "NO_DELIVERED"
            bucket = "NO_DELIVERED"
        else:
            first_del = delivered_rows[time_col].min()
            last_del = delivered_rows[time_col].max()
            minutes_since_last = (now_et_naive - last_del).total_seconds() / 60
            hours_since_first = (now_et_naive - first_del).total_seconds() / 3600
            per_hour = (delivered_cnt / hours_since_first) if hours_since_first and hours_since_first > 0 else None

            status_flag = "HAS_DELIVERED"
            if minutes_since_last > 60:
                bucket = "RED"
            elif minutes_since_last > 30:
                bucket = "YELLOW"
            else:
                bucket = "OK"

        rows.append({
            "Route": route,
            "DriverName": driver,
            "FleeName": flee,
            "Total": total,
            "Success(Delivered)": delivered_cnt,
            "Failed(*FAIL*)": failed_cnt,
            "Remaining": remaining,
            "CompletionRate": completion,
            "1stDeliveryTime": first_del,
            "HoursSinceFirstDelivery": round(hours_since_first, 2) if hours_since_first is not None else None,
            "DeliveriesPerHour": round(per_hour, 2) if per_hour is not None else None,
            "LatestDeliveredTime": last_del,
            "MinutesSinceLast": round(minutes_since_last, 1) if minutes_since_last is not None else None,
            "StatusFlag": status_flag,
            "AlertBucket": bucket
        })

    route_df = pd.DataFrame(rows)

    # sort: NO_DELIVERED first; then stalled first
    route_df["_sort"] = route_df["MinutesSinceLast"].fillna(10**9)
    route_df.sort_values(["StatusFlag", "_sort"], ascending=[True, False], inplace=True)
    route_df.drop(columns="_sort", inplace=True)

    # Summary
    sum_df = route_df.copy()
    sum_df["FleeName"] = sum_df["FleeName"].fillna("UNKNOWN")
    summary_df = sum_df.groupby("FleeName").agg(
        Routes=("Route", "nunique"),
        TotalPkgs=("Total", "sum"),
        Delivered=("Success(Delivered)", "sum"),
        Failed=("Failed(*FAIL*)", "sum"),
        Remaining=("Remaining", "sum"),
        NoDeliveredRoutes=("StatusFlag", lambda s: int((s == "NO_DELIVERED").sum())),
        RedRoutes=("AlertBucket", lambda s: int((s == "RED").sum())),
        YellowRoutes=("AlertBucket", lambda s: int((s == "YELLOW").sum())),
        AvgDeliveriesPerHour=("DeliveriesPerHour", "mean"),
    ).reset_index()
    summary_df["CompletionRate"] = (summary_df["Delivered"] / summary_df["TotalPkgs"]).fillna(0.0)
    summary_df["AvgDeliveriesPerHour"] = summary_df["AvgDeliveriesPerHour"].round(2)

    # Exceptionsï¼ˆå›ºå®šè§„åˆ™ï¼‰
    exc_df = route_df[
        (route_df["StatusFlag"] == "NO_DELIVERED") |
        ((route_df["MinutesSinceLast"].fillna(0) > 120) & (route_df["Remaining"] > 0)) |
        ((route_df["DeliveriesPerHour"].fillna(999) < 10) & (route_df["Remaining"] > 0))
    ].copy()

    # time checks
    after_3pm = now_et.hour >= 15
    after_6pm = now_et.hour >= 18
    check_3pm = route_df[route_df["CompletionRate"] < 0.5].copy() if after_3pm else route_df.iloc[0:0].copy()
    check_6pm = route_df[route_df["CompletionRate"] < 0.8].copy() if after_6pm else route_df.iloc[0:0].copy()

    # Excel output
    wb = Workbook()

    route_cols = [
        "Route","DriverName","FleeName","Total","Success(Delivered)","Failed(*FAIL*)","Remaining","CompletionRate",
        "1stDeliveryTime","HoursSinceFirstDelivery","DeliveriesPerHour",
        "LatestDeliveredTime","MinutesSinceLast","StatusFlag","AlertBucket"
    ]

    # RouteMonitor
    ws1 = wb.active
    ws1.title = "RouteMonitor"
    ws1.append(route_cols)
    for r in route_df[route_cols].itertuples(index=False):
        ws1.append(list(r))
    style_header(ws1)
    apply_route_colors(ws1, route_cols)
    autosize(ws1)

    # Summary
    ws2 = wb.create_sheet("Summary")
    summary_cols = ["FleeName","Routes","TotalPkgs","Delivered","Failed","Remaining","CompletionRate",
                    "NoDeliveredRoutes","RedRoutes","YellowRoutes","AvgDeliveriesPerHour"]
    ws2.append(summary_cols)
    for r in summary_df[summary_cols].itertuples(index=False):
        ws2.append(list(r))
    style_header(ws2)
    cr_idx = summary_cols.index("CompletionRate") + 1
    for rr in range(2, ws2.max_row + 1):
        ws2.cell(rr, cr_idx).number_format = "0.00%"
    autosize(ws2)

    # Exceptions
    ws3 = wb.create_sheet("Exceptions")
    ws3.append(route_cols)
    for r in exc_df[route_cols].itertuples(index=False):
        ws3.append(list(r))
    style_header(ws3)
    apply_route_colors(ws3, route_cols)
    autosize(ws3)

    # 3pm check
    ws4 = wb.create_sheet("3pm check")
    ws4["A1"] = "RunTime (ET)"; ws4["B1"] = now_et.strftime("%Y-%m-%d %H:%M:%S")
    ws4["A2"] = "Rule"; ws4["B2"] = "At/after 3:00 PM ET: CompletionRate < 50%"
    ws4["A3"] = "RuleApplied"; ws4["B3"] = "YES" if after_3pm else "NO (before 3 PM ET)"
    ws4.append([]); ws4.append(route_cols)
    for r in check_3pm[route_cols].itertuples(index=False):
        ws4.append(list(r))
    style_header(ws4)
    light_orange = PatternFill("solid", fgColor="FCE4D6")
    for rr in range(6, ws4.max_row + 1):
        for cc in range(1, ws4.max_column + 1):
            ws4.cell(rr, cc).fill = light_orange
    autosize(ws4)

    # 6pm check
    ws5 = wb.create_sheet("6pm check")
    ws5["A1"] = "RunTime (ET)"; ws5["B1"] = now_et.strftime("%Y-%m-%d %H:%M:%S")
    ws5["A2"] = "Rule"; ws5["B2"] = "At/after 6:00 PM ET: CompletionRate < 80%"
    ws5["A3"] = "RuleApplied"; ws5["B3"] = "YES" if after_6pm else "NO (before 6 PM ET)"
    ws5.append([]); ws5.append(route_cols)
    for r in check_6pm[route_cols].itertuples(index=False):
        ws5.append(list(r))
    style_header(ws5)
    deep_orange = PatternFill("solid", fgColor="F4B084")
    for rr in range(6, ws5.max_row + 1):
        for cc in range(1, ws5.max_column + 1):
            ws5.cell(rr, cc).fill = deep_orange
    autosize(ws5)

    # Meta
    ws6 = wb.create_sheet("Meta")
    ws6["A1"] = "Now (ET) used for calculation"
    ws6["B1"] = now_et.strftime("%Y-%m-%d %H:%M:%S")
    ws6["A3"] = "Color rules"
    ws6["A4"] = "Purple: NO_DELIVERED"
    ws6["A5"] = "Yellow: MinutesSinceLast > 30 and <= 60"
    ws6["A6"] = "Red: MinutesSinceLast > 60"
    ws6["A8"] = "Exceptions criteria"
    ws6["A9"] = "1) NO_DELIVERED OR 2) MinutesSinceLast>120 & Remaining>0 OR 3) DeliveriesPerHour<10 & Remaining>0"
    ws6["A11"] = "3pm rule"
    ws6["A12"] = "At/after 3:00 PM ET: CompletionRate < 50%"
    ws6["A14"] = "6pm rule"
    ws6["A15"] = "At/after 6:00 PM ET: CompletionRate < 80%"
    autosize(ws6, cap=70)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------- Streamlit UI ----------------
st.set_page_config(page_title="Daily DSP Operation Check", layout="wide")
st.title("Daily DSP Operation Check â€” æ¯å¤©ä¸€é”®ç”Ÿæˆ Excel")

with st.expander("è¾“å…¥è¦æ±‚ï¼ˆå¿…é¡»ï¼‰", expanded=True):
    st.markdown(
        """
- åŽŸå§‹ Excel å¿…é¡»åŒ…å«ï¼š
  - **Båˆ—** = Route
  - **Jåˆ—** = åŒ…è£¹çŠ¶æ€ï¼ˆStatusï¼‰
  - **Låˆ—** = çŠ¶æ€æ—¶é—´ï¼ˆStatusTimeï¼‰
- å¯é€‰åˆ—ï¼ˆæœ‰å°±è‡ªåŠ¨å¸¦ä¸Šï¼‰ï¼š
  - åˆ—ååŒ…å« **Flee** â†’ ä½œä¸º FleeName
  - åˆ—ååŒ…å« **Driver** â†’ ä½œä¸º DriverName
        """
    )

# âœ… è¿™å°±æ˜¯ä½ è¦çš„â€œç½‘é¡µä¸­ä¸Šä¼ æ–‡ä»¶çš„ä½ç½®â€
uploaded = st.file_uploader("ðŸ“¤ Upload your .xlsx file here", type=["xlsx"])

if uploaded:
    try:
        raw_df = pd.read_excel(uploaded, engine="openpyxl", dtype=str)
    except Exception:
        raw_df = pd.read_excel(uploaded, dtype=str)

    output_bytes = build_excel_bytes(raw_df)
    ts = datetime.datetime.now(pytz.timezone(TZ)).strftime("%Y%m%d_%H%M%S")

    st.success("âœ… å·²ç”ŸæˆæŠ¥è¡¨ï¼Œç‚¹å‡»ä¸‹è½½ï¼š")
    st.download_button(
        label="â¬‡ï¸ Download Excel Report",
        data=output_bytes,
        file_name=f"route_monitor_{ts}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
else:
    st.info("è¯·å…ˆä¸Šä¼  Excel æ–‡ä»¶ã€‚")
